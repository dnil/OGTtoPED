import argparse
import sys

from openpyxl import load_workbook

import config
from sample import Sample
from family import update_family, family_ped

# CLI

parser = argparse.ArgumentParser(description="Convert OGT xlsx to PED file")

parser.add_argument("orderform", 
    help="OGT order form with sample ID, status and family groups.")
parser.add_argument("outfile", help="Output PED file", nargs='?')
parser.add_argument("-D", "--debug", help="Enable DEBUG output.", 
                    action="store_true")

args = parser.parse_args()

if config.debug:
    print(sys.stderr, "DEBUG output turned on.")
    config.debug = True

config.outfile = args.outfile

# Truncate output ped file

if config.outfile is not None:
    out = open(config.outfile, 'w')
else:
    out = sys.stdout

# Open workbook

wb = load_workbook(filename = args.orderform)
ws = wb.active

# Sanity checks

if ws.title != "material form": 
    print(sys.stderr, "WARNING: Non standard active sheet name ", ws.title)

if (ws['B12'].value != "Customer Sample ID" 
        or ws['M12'].value != "Additional Experimental Comments" 
        or ws['C12'].value != "Source (cells, tissue etc)"):
    print(sys.stderr, ("Unexpected table / cell layout: check to see"
        "that sheet is ok, and ask to have the script updated."))

    exit(1)

# Main
# Iterate over all rows, parse row blocks 
in_sample_section = False
in_family = False

max_rows = 1024

samples_found = 0
family = []
family_count = 0 

for rownum in range(1,max_rows+1):
    cell=ws["B" + str(rownum)]

    if not in_sample_section:
        if cell.value == "Customer Sample ID":
            if config.debug:
                print(sys.stderr, "Found sample ID tag.")
            in_sample_section = True
    else:
        if cell.value is not None:
            # Found a new sample row.
            sample_id = cell.value
            sample_id.rstrip()

            if not in_family:
                if config.debug:
                    print(sys.stderr, ("New family, starting with sample "
                                         "'{}'").format(sample_id))
                family_count += 1

            info_cell = ws["M" + str(rownum)]
            info = info_cell.value
            if info is None:
                info = "NA"
            info.rstrip()

            tissue_cell =  ws["C" + str(rownum)]
            tissue = tissue_cell.value
            if tissue is None:
                tissue = "NA"
            tissue.rstrip()

            sample = Sample(sample_id, info, tissue)
            in_family = True
            family.append(sample)
            
            if sample.info.find("singleton") != -1:          
                # Found a singleton!
                sample.affected = True
                update_family(family)
                print >> out, family_ped(family, family_count).rstrip()
                # This ends the current family.
                if config.debug:
                    print(sys.stderr, "Found a singleton. Family complete.")
                family = []
                in_family = False
                # Note that the next row may be a None or a new family member..

            samples_found += 1

        elif cell.value is None: 
            # Value None means an empty row.
            if in_family:
                # This ends the current family.
                if config.debug:
                    print(sys.stderr, "Family complete.")

                update_family(family)
                print >> out, family_ped(family, family_count).rstrip()

                family = []
                in_family = False
